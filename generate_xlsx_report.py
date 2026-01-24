#!/usr/bin/env python3
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Test categories
CATEGORIES = [
    (
        "CRUD",
        [
            "insert_single",
            "insert_bulk_100",
            "select_pk",
            "select_filter",
            "update_single",
            "update_bulk",
            "delete_single",
        ],
    ),
    (
        "Queries",
        [
            "filter_simple",
            "filter_complex",
            "filter_in",
            "order_limit",
            "aggregate_count",
            "aggregate_mixed",
        ],
    ),
    (
        "Relations",
        [
            "join_simple",
            "join_filter",
            "prefetch_related",
            "nested_prefetch",
        ],
    ),
    (
        "Concurrent",
        [
            "concurrent_10",
            "concurrent_25",
            "concurrent_50",
            "concurrent_75",
            "concurrent_100",
            "concurrent_150",
            "concurrent_200",
        ],
    ),
]

# Sheets to create with their metrics
SHEETS = [
    ("ops_sec", "ops_per_sec", "ops/sec", "higher"),
    ("Peak Memory", "memory_peak_mb", "MB", "lower"),
    ("CPU Time", "cpu_user_sec", "sec", "lower"),
    ("Mean Latency", "mean_ms", "ms", "lower"),
    ("P95 Latency", "p95_ms", "ms", "lower"),
    ("P99 Latency", "p99_ms", "ms", "lower"),
]

# Raw drivers to exclude from charts (keep in tables as baseline)
EXCLUDE_FROM_CHARTS = ["asyncpg", "aiomysql", "aiosqlite"]


def create_header_style():
    return {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def create_cell_style():
    return {
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def get_metric_value(test_data, metric_key):
    """Get metric value, handling special cases like CPU time."""
    if "error" in test_data:
        return None

    if metric_key == "cpu_user_sec":
        user = test_data.get("cpu_user_sec", 0) or 0
        system = test_data.get("cpu_system_sec", 0) or 0
        return user + system

    return test_data.get(metric_key)


def create_category_table(
    ws, start_row, category_name, tests, orms, results, metric_key, metric_unit
):
    """Create a table for one category. Returns the row after the table."""
    header_style = create_header_style()
    cell_style = create_cell_style()

    # Category title
    ws.cell(row=start_row, column=1, value=category_name)
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    # Header row
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Test")
    apply_style(ws.cell(row=header_row, column=1), header_style)

    for col_idx, orm in enumerate(orms, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=orm)
        apply_style(cell, header_style)

    # Data rows
    data_start_row = header_row + 1
    for row_offset, test in enumerate(tests):
        row_idx = data_start_row + row_offset

        # Test name
        test_cell = ws.cell(row=row_idx, column=1, value=test)
        test_cell.font = Font(bold=True)
        apply_style(test_cell, cell_style)

        # ORM values
        for col_idx, orm in enumerate(orms, start=2):
            value = None
            if orm in results and test in results[orm]:
                value = get_metric_value(results[orm][test], metric_key)

            cell = ws.cell(
                row=row_idx, column=col_idx, value=value if value is not None else "N/A"
            )
            apply_style(cell, cell_style)
            if isinstance(value, (int, float)):
                if metric_unit == "ops/sec":
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"

    data_end_row = data_start_row + len(tests) - 1

    return header_row, data_start_row, data_end_row


def create_category_chart(
    ws,
    category_name,
    header_row,
    data_start_row,
    data_end_row,
    orms,
    chart_col,
    metric_unit,
    better,
    results,
    tests,
    metric_key,
):
    """Create a bar chart for one category (excluding baseline drivers and N/A ORMs)."""
    # Filter out excluded ORMs and those with all N/A values
    chart_orms = []
    for orm in orms:
        if orm in EXCLUDE_FROM_CHARTS:
            continue
        # Check if ORM has at least one valid value for these tests
        has_valid = False
        for test in tests:
            if orm in results and test in results[orm]:
                value = get_metric_value(results[orm][test], metric_key)
                if value is not None:
                    has_valid = True
                    break
        if has_valid:
            chart_orms.append(orm)

    if not chart_orms:
        return

    # Find column indices for chart ORMs
    chart_col_indices = []
    for orm in chart_orms:
        if orm in orms:
            chart_col_indices.append(orms.index(orm) + 2)  # +2 because col 1 is Test

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"{category_name} ({metric_unit})"
    chart.y_axis.title = metric_unit
    chart.width = 15
    chart.height = 10

    # Add data series for each ORM (excluding baseline)
    cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

    for col_idx in chart_col_indices:
        data = Reference(ws, min_col=col_idx, min_row=header_row, max_row=data_end_row)
        chart.add_data(data, titles_from_data=True)

    chart.set_categories(cats)

    # Place chart to the right of table
    chart_cell = f"{get_column_letter(chart_col)}{header_row - 1}"
    ws.add_chart(chart, chart_cell)


def create_sheet_with_categories(
    wb, sheet_name, orms, results, metric_key, metric_unit, better
):
    """Create a sheet with tables and charts for each category."""
    ws = wb.create_sheet(title=sheet_name)

    better_text = "↑ higher better" if better == "higher" else "↓ lower better"

    # Title
    ws.cell(row=1, column=1, value=f"{sheet_name} ({metric_unit}) — {better_text}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, len(orms) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    # Chart column (after ORM columns + some space)
    chart_col = len(orms) + 3

    current_row = 3
    for category_name, tests in CATEGORIES:
        header_row, data_start_row, data_end_row = create_category_table(
            ws,
            current_row,
            category_name,
            tests,
            orms,
            results,
            metric_key,
            metric_unit,
        )

        create_category_chart(
            ws,
            category_name,
            header_row,
            data_start_row,
            data_end_row,
            orms,
            chart_col,
            metric_unit,
            better,
            results,
            tests,
            metric_key,
        )

        # Next category starts after table + some space
        current_row = data_end_row + 3

    return ws


def create_scalability_sheet(wb, orms, results):
    """Create a sheet showing latency scalability under concurrent load."""
    ws = wb.create_sheet(title="Scalability")

    header_style = create_header_style()
    cell_style = create_cell_style()

    # Title
    ws.cell(
        row=1,
        column=1,
        value="Scalability — Mean Latency vs Concurrency (↓ lower better)",
    )
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    concurrent_tests = [
        "concurrent_10",
        "concurrent_25",
        "concurrent_50",
        "concurrent_75",
        "concurrent_100",
        "concurrent_150",
        "concurrent_200",
    ]
    concurrency_values = [10, 25, 50, 75, 100, 150, 200]

    # Filter ORMs (exclude raw drivers and those with all N/A)
    chart_orms = []
    for orm in orms:
        if orm in EXCLUDE_FROM_CHARTS:
            continue
        has_valid = False
        for test in concurrent_tests:
            if orm in results and test in results[orm]:
                value = get_metric_value(results[orm][test], "mean_ms")
                if value is not None:
                    has_valid = True
                    break
        if has_valid:
            chart_orms.append(orm)

    if not chart_orms:
        ws.cell(row=3, column=1, value="No valid data for concurrent tests")
        return ws

    # Table header
    ws.cell(row=3, column=1, value="Concurrency")
    apply_style(ws.cell(row=3, column=1), header_style)

    for col_idx, orm in enumerate(chart_orms, start=2):
        cell = ws.cell(row=3, column=col_idx, value=orm)
        apply_style(cell, header_style)

    # Data rows
    for row_offset, (test, concurrency) in enumerate(
        zip(concurrent_tests, concurrency_values)
    ):
        row_idx = 4 + row_offset

        ws.cell(row=row_idx, column=1, value=concurrency)
        apply_style(ws.cell(row=row_idx, column=1), cell_style)

        for col_idx, orm in enumerate(chart_orms, start=2):
            value = None
            if orm in results and test in results[orm]:
                test_data = results[orm][test]
                if "error" not in test_data:
                    value = test_data.get("mean_ms", 0)

            cell = ws.cell(
                row=row_idx, column=col_idx, value=value if value is not None else "N/A"
            )
            apply_style(cell, cell_style)
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    data_end_row = 4 + len(concurrent_tests) - 1

    # Create line chart
    chart = LineChart()
    chart.title = "Latency Scalability (Mean ms)"
    chart.y_axis.title = "Latency (ms)"
    chart.x_axis.title = "Concurrency"
    chart.width = 20
    chart.height = 12
    chart.style = 10

    # Add data series for each ORM
    data = Reference(
        ws, min_col=2, min_row=3, max_col=len(chart_orms) + 1, max_row=data_end_row
    )
    cats = Reference(ws, min_col=1, min_row=4, max_row=data_end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Make lines thinner
    for series in chart.series:
        series.graphicalProperties.line.width = 20000  # EMUs (thin line ~0.5pt)

    # Place chart
    ws.add_chart(chart, "A8")

    # Column widths
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(chart_orms) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    return ws


def generate_report(results_path: Path, output_path: Path):
    """Generate XLSX report from benchmark results."""
    with open(results_path) as f:
        data = json.load(f)

    results = data["results"]
    orms = list(results.keys())

    wb = Workbook()
    wb.remove(wb.active)

    # Create sheets
    for sheet_name, metric_key, metric_unit, better in SHEETS:
        create_sheet_with_categories(
            wb, sheet_name, orms, results, metric_key, metric_unit, better
        )

    # Create scalability sheet
    create_scalability_sheet(wb, orms, results)

    wb.save(output_path)
    print(f"Report saved to: {output_path}")


def process_all_results(results_dir: Path, output_dir: Path = None):
    """Process all result directories and generate xlsx reports."""
    if output_dir is None:
        output_dir = Path.cwd()

    result_dirs = (
        sorted(results_dir.glob("*_postgres"))
        + sorted(results_dir.glob("*_mysql"))
        + sorted(results_dir.glob("*_sqlite"))
    )

    if not result_dirs:
        print(f"No result directories found in {results_dir}")
        return

    for result_dir in result_dirs:
        # Find json file
        json_files = list(result_dir.glob("*.json"))
        json_files = [f for f in json_files if f.name != "env.json"]

        if not json_files:
            print(f"No results JSON found in {result_dir}")
            continue

        results_path = json_files[0]
        # Use directory name as xlsx filename
        output_path = output_dir / f"{result_dir.name}.xlsx"

        print(
            f"Processing: {result_dir.name}/{results_path.name} -> {output_path.name}"
        )
        try:
            generate_report(results_path, output_path)
        except Exception as e:
            print(f"  Error: {e}")


def main():
    results_dir = Path("/home/nikita/Workspace/oxyde-benchmarks/results")

    if len(sys.argv) < 2:
        # Process all results
        process_all_results(results_dir)
    elif sys.argv[1] == "--all":
        process_all_results(results_dir)
    else:
        # Single file mode
        results_path = Path(sys.argv[1])
        if len(sys.argv) >= 3:
            output_path = Path(sys.argv[2])
        else:
            output_path = results_path.with_suffix(".xlsx")
        generate_report(results_path, output_path)


if __name__ == "__main__":
    main()
